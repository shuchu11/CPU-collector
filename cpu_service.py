"""
cpu_service.py — Standalone CPU Monitoring Microservice
Features:
  - Background continuous monitoring (start/stop)
  - Real-time append to Excel (one row per second, one col per core)
  - Plot from Excel data (heatmap + timeseries)
"""

import io
import os
import re
import time
import subprocess
import threading
from abc import ABC, abstractmethod
from typing import Dict, List, Any, Optional
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from openpyxl import load_workbook, Workbook

from flask import Flask, jsonify, request, send_file

app = Flask(__name__)
PORT = 5001


# ============================================================================
# UTILS
# ============================================================================

def run_cmd(cmd: str, timeout: int = 30) -> Optional[str]:
    try:
        result = subprocess.run(
            cmd, shell=True, capture_output=True, text=True, timeout=timeout
        )
        return result.stdout.strip() if result.returncode == 0 else None
    except Exception:
        return None


def nsenter(cmd: str, timeout: int = 30) -> Optional[str]:
    return run_cmd(f"nsenter -t 1 -m -u -n -i {cmd}", timeout)


# ============================================================================
# BASE COLLECTOR
# ============================================================================

class BaseCollector(ABC):
    def __init__(self):
        self.samples = []
        self.timestamps = []

    @abstractmethod
    def collect_sample(self) -> Optional[Dict[str, Any]]:
        pass

    def gather(self, duration: int, include_timeseries: bool = False) -> Dict[str, Any]:
        self.samples, self.timestamps = [], []
        for _ in range(duration):
            ts = time.time()
            sample = self.collect_sample()
            if sample is not None:
                self.timestamps.append(ts)
                self.samples.append(sample)
            time.sleep(1)
        return self.aggregate(include_timeseries)

    @abstractmethod
    def aggregate(self, include_timeseries: bool = True) -> Dict[str, Any]:
        pass

    def get_stats(self, values: List[float]) -> Dict[str, float]:
        if not values:
            return {"min": 0, "max": 0, "avg": 0}
        return {
            "min": round(min(values), 2),
            "max": round(max(values), 2),
            "avg": round(sum(values) / len(values), 2),
        }


# ============================================================================
# CPU COLLECTORS
# ============================================================================

class CPUCollector(BaseCollector):
    def __init__(self, include_breakdown: bool = False):
        super().__init__()
        self.prev_stats = None
        self.include_breakdown = include_breakdown

    def _read_cpu_stats(self) -> Dict[str, Dict[str, int]]:
        out = nsenter("cat /proc/stat")
        if not out:
            return {}
        cpu_stats = {}
        for line in out.split("\n"):
            if not line.startswith("cpu"):
                continue
            parts = line.split()
            if len(parts) < 8:
                continue
            values = [int(x) for x in parts[1:8]]
            cpu_stats[parts[0]] = {
                "user": values[0], "nice": values[1], "system": values[2],
                "idle": values[3], "iowait": values[4], "irq": values[5],
                "softirq": values[6], "total": sum(values),
            }
        return cpu_stats

    def collect_sample(self) -> Optional[Dict[str, Any]]:
        current_stats = self._read_cpu_stats()
        if not current_stats:
            return None
        sample = {}
        if self.prev_stats:
            for cpu_name, current in current_stats.items():
                if cpu_name not in self.prev_stats:
                    continue
                prev = self.prev_stats[cpu_name]
                delta_total = current["total"] - prev["total"]
                if delta_total == 0:
                    sample[cpu_name] = {"usage": 0.0}
                    continue
                delta_idle = current["idle"] - prev["idle"]
                sample[cpu_name] = {
                    "usage": round(((delta_total - delta_idle) / delta_total) * 100, 2)
                }
                if self.include_breakdown:
                    sample[cpu_name]["breakdown"] = {
                        k: round(((current[k] - prev[k]) / delta_total) * 100, 2)
                        for k in ("user", "system", "iowait", "irq", "softirq")
                    }
        self.prev_stats = current_stats
        return sample if sample else None

    def aggregate(self, include_timeseries: bool = True) -> Dict[str, Any]:
        if not self.samples:
            return {"error": "no samples collected"}
        cpu_data: Dict[str, list] = {}
        for sample in self.samples:
            for cpu_name, data in sample.items():
                cpu_data.setdefault(cpu_name, []).append(data["usage"])

        timestamps_rounded = [round(ts, 2) for ts in self.timestamps]
        result = {"duration": len(self.timestamps), "samples": len(self.samples), "cpus": {}}
        for cpu_name, usages in cpu_data.items():
            stats = self.get_stats(usages)
            if include_timeseries:
                stats["timeseries"] = {"timestamps": timestamps_rounded, "percent": usages}
            result["cpus"][cpu_name] = {"usage": stats}

        online_cpus = run_cmd("cat /sys/devices/system/cpu/online") or ""
        numbers = re.findall(r"\d+", online_cpus)
        system_max_cpu = max(int(n) for n in numbers) if numbers else 0
        offline_cpus = [
            i for i in range(system_max_cpu + 1)
            if run_cmd(f"cat /sys/devices/system/cpu/cpu{i}/online 2>/dev/null") == "0"
        ]
        result["system_max_cpu"] = system_max_cpu
        result["offline_cpus"] = offline_cpus
        result["isolated_cpus"] = run_cmd("cat /sys/devices/system/cpu/isolated")
        return result


class ContextSwitchCollector(BaseCollector):
    def __init__(self, cpu_filter: Optional[List[int]] = None):
        super().__init__()
        self.cpu_filter = cpu_filter
        self.prev_ctxt: Dict[str, int] = {}

    def collect_sample(self) -> Optional[Dict[str, Any]]:
        out = nsenter("cat /proc/stat")
        if not out:
            return None
        sample: Dict[str, Any] = {}
        for line in out.split("\n"):
            if line.startswith("ctxt"):
                sample["total"] = int(line.split()[1])
        if self.cpu_filter:
            for cpu_id in self.cpu_filter:
                schedstat = nsenter(f"cat /proc/schedstat | grep 'cpu{cpu_id} ' 2>/dev/null")
                if schedstat:
                    parts = schedstat.split()
                    if len(parts) >= 8:
                        sample[f"cpu{cpu_id}"] = int(parts[7])
        if self.prev_ctxt:
            for key, value in list(sample.items()):
                if key in self.prev_ctxt:
                    sample[f"{key}_rate"] = value - self.prev_ctxt[key]
        self.prev_ctxt = dict(sample)
        return sample

    def aggregate(self, include_timeseries: bool = True) -> Dict[str, Any]:
        if not self.samples:
            return {"error": "no samples collected"}
        ctxt_data: Dict[str, list] = {}
        for sample in self.samples:
            for key, value in sample.items():
                ctxt_data.setdefault(key, []).append(value)
        timestamps_rounded = [round(ts, 2) for ts in self.timestamps]
        result = {"duration": len(self.timestamps), "samples": len(self.samples), "context_switches": {}}
        for key, values in ctxt_data.items():
            stats = self.get_stats(values)
            if include_timeseries:
                stats["timeseries"] = {"timestamps": timestamps_rounded, "count": values}
            result["context_switches"][key] = stats
        return result


class CPUGovernorCollector:
    @staticmethod
    def get_governor() -> Dict[str, Any]:
        out = run_cmd(
            "nsenter -t 1 -m -u -n -i bash -c "
            "'cat /sys/devices/system/cpu/cpu*/cpufreq/scaling_governor | sort -u'"
        )
        return {"governors": out.split("\n") if out else None}


class CPUIdleCollector:
    @staticmethod
    def get_idle_states() -> Dict[str, Any]:
        out = run_cmd(
            "nsenter -t 1 -m -u -n -i bash -c "
            "'cat /sys/devices/system/cpu/cpu0/cpuidle/state*/disable'"
        )
        return {"idle_states_disabled": out.split("\n") if out else None}


class IRQAffinityCollector:
    @staticmethod
    def get_affinity(pattern: str = "ens") -> Dict[str, Any]:
        out = run_cmd(
            f"nsenter -t 1 -m -u -n -i bash -c "
            f"'grep -H . /proc/irq/*/smp_affinity_list | grep {pattern}'"
        )
        if not out:
            return {"irq_affinity": {}}
        affinities = {}
        for line in out.split("\n"):
            if ":" in line:
                parts = line.split(":")
                irq_num = parts[0].split("/")[3]
                affinities[irq_num] = parts[1]
        return {"irq_affinity": affinities}




# ============================================================================
# THREAD CPU COLLECTOR
# ============================================================================

class ThreadCPUCollector(BaseCollector):
    """Monitor per-thread CPU usage with core affinity tracking"""

    def __init__(self, pid: Optional[int] = None, pgrep: str = "nr-softmodem"):
        super().__init__()
        self.pid = pid
        self.pgrep = pgrep
        self.cached_pid = None

    def _discover_pid(self) -> Optional[int]:
        if self.cached_pid:
            return self.cached_pid
        if self.pid:
            self.cached_pid = self.pid
            return self.pid
        pids = nsenter(f"pgrep {self.pgrep}")
        if not pids:
            return None
        pid_list = [int(p) for p in pids.split("\n") if p.strip()]
        if not pid_list:
            return None
        self.cached_pid = min(pid_list)
        return self.cached_pid

    def _read_thread_stat(self, pid: int, tid: int):
        """Read /proc/<pid>/task/<tid>/stat — returns (utime, stime, cutime, cstime, processor)"""
        stat = nsenter(f"cat /proc/{pid}/task/{tid}/stat 2>/dev/null")
        if not stat:
            return None
        # comm may contain spaces and is wrapped in parens; split after last ')'
        rp = stat.rfind(")")
        if rp == -1:
            return None
        fields = stat[rp + 2:].split()
        # fields[0] = state, fields[11] = utime, fields[12] = stime,
        # fields[13] = cutime, fields[14] = cstime, fields[36] = processor
        try:
            utime  = int(fields[11])
            stime  = int(fields[12])
            processor = int(fields[36])
            return utime, stime, processor
        except (IndexError, ValueError):
            return None

    def _read_thread_comm(self, pid: int, tid: int) -> str:
        comm = nsenter(f"cat /proc/{pid}/task/{tid}/comm 2>/dev/null")
        return comm.strip() if comm else str(tid)

    def collect_sample(self) -> Optional[Dict[str, Any]]:
        pid = self._discover_pid()
        if not pid:
            return None

        # List all TIDs
        task_list = nsenter(f"ls /proc/{pid}/task 2>/dev/null")
        if not task_list:
            return None
        tids = [int(t) for t in task_list.split() if t.strip().isdigit()]

        # Read current stats
        current: Dict[int, Any] = {}
        for tid in tids:
            s = self._read_thread_stat(pid, tid)
            if s:
                current[tid] = s  # (utime, stime, processor)

        sample = {}
        if hasattr(self, "_prev_thread_stats") and self._prev_thread_stats:
            hz = 100  # USER_HZ, standard on Linux
            for tid, (utime, stime, processor) in current.items():
                if tid not in self._prev_thread_stats:
                    continue
                prev_utime, prev_stime, _ = self._prev_thread_stats[tid]
                delta = (utime - prev_utime) + (stime - prev_stime)
                cpu_pct = round((delta / hz) * 100, 2)
                comm = self._read_thread_comm(pid, tid)
                sample[str(tid)] = {"name": comm, "cpu": cpu_pct, "core": processor}

        self._prev_thread_stats = current
        return sample if sample else None

    def aggregate(self, include_timeseries: bool = True) -> Dict[str, Any]:
        if not self.samples:
            return {"error": "no samples collected"}
        from collections import Counter
        thread_stats: Dict[str, Any] = {}
        for sample in self.samples:
            for tid, data in sample.items():
                if tid not in thread_stats:
                    thread_stats[tid] = {"name": data["name"], "cpu_samples": [], "core_samples": []}
                thread_stats[tid]["cpu_samples"].append(data["cpu"])
                thread_stats[tid]["core_samples"].append(data["core"])

        timestamps_rounded = [round(ts, 2) for ts in self.timestamps]
        result = {
            "pid": self._discover_pid(),
            "duration": len(self.timestamps),
            "total_threads": len(thread_stats),
            "threads": [],
        }
        for tid, data in thread_stats.items():
            cpu_values = data["cpu_samples"]
            core_values = data["core_samples"]
            core_distribution = dict(Counter(core_values))
            cpu_stats = self.get_stats(cpu_values)
            thread_info = {
                "tid": tid,
                "name": data["name"],
                "cpu_usage": cpu_stats,
                "core_distribution": core_distribution,
                "primary_core": max(core_distribution, key=core_distribution.get),
            }
            if include_timeseries:
                thread_info["cpu_usage"]["timeseries"] = {
                    "timestamps": timestamps_rounded,
                    "percent": cpu_values,
                }
                thread_info["core_timeseries"] = {
                    "timestamps": timestamps_rounded,
                    "cores": core_values,
                }
            result["threads"].append(thread_info)

        result["threads"].sort(key=lambda x: x["cpu_usage"]["avg"], reverse=True)
        online_cpus = run_cmd("cat /sys/devices/system/cpu/online")
        if online_cpus:
            numbers = re.findall(r"\d+", online_cpus)
            result["system_max_cpu"] = max(int(n) for n in numbers) if numbers else 0
        offline_cpus = []
        system_max_cpu = result.get("system_max_cpu", 0)
        for cpu_id in range(system_max_cpu + 1):
            status = run_cmd(f"cat /sys/devices/system/cpu/cpu{cpu_id}/online 2>/dev/null")
            if status == "0":
                offline_cpus.append(cpu_id)
        result["offline_cpus"] = offline_cpus
        return result


# ============================================================================
# THREAD BACKGROUND MONITOR — writes to Excel
# ============================================================================

class ThreadBackgroundMonitor:
    """
    Background thread that samples ThreadCPUCollector every N seconds
    and appends aggregated results to an Excel file.

    Excel layout (sheet: "thread_cpu"):
        timestamp | tid | name | avg_cpu | min_cpu | max_cpu | primary_core | core_0 | core_1 | ...
    """

    def __init__(self):
        self._thread: Optional[threading.Thread] = None
        self._stop_event = threading.Event()
        self._lock = threading.Lock()
        self._write_lock = threading.Lock()
        self.xlsx_path: Optional[str] = None
        self.running = False
        self.rows_written = 0
        self.started_at: Optional[str] = None
        self._header_written = False
        self._max_core = 0

    def start(self, xlsx_path: str, pgrep: str = "nr-softmodem",
              sample_interval: int = 10) -> Dict[str, Any]:
        with self._lock:
            if self.running:
                return {"status": "already_running", "xlsx": self.xlsx_path}
            self.xlsx_path = xlsx_path
            self._pgrep = pgrep
            self._sample_interval = sample_interval
            self._stop_event.clear()
            self.rows_written = 0
            self.started_at = datetime.now().isoformat(timespec="seconds")
            self._header_written = False
            self._max_core = 0
            os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "thread_cpu"
            wb.save(xlsx_path)
            self._thread = threading.Thread(target=self._run, daemon=True)
            self._thread.start()
            self.running = True
        return {"status": "started", "xlsx": xlsx_path, "started_at": self.started_at,
                "pgrep": pgrep, "sample_interval": sample_interval}

    def stop(self) -> Dict[str, Any]:
        with self._lock:
            if not self.running:
                return {"status": "not_running"}
            self._stop_event.set()
        if self._thread:
            self._thread.join(timeout=self._sample_interval + 5)
        with self._lock:
            self.running = False
        return {
            "status": "stopped",
            "xlsx": self.xlsx_path,
            "rows_written": self.rows_written,
            "started_at": self.started_at,
            "stopped_at": datetime.now().isoformat(timespec="seconds"),
        }

    def status(self) -> Dict[str, Any]:
        with self._lock:
            return {
                "running": self.running,
                "xlsx": self.xlsx_path,
                "rows_written": self.rows_written,
                "started_at": self.started_at,
            }

    def _append_batch(self, ts: str, threads: List[Dict[str, Any]]):
        """Append one row per thread for this sample period"""
        # Find max core across all threads
        for t in threads:
            for core_key in t.get("core_distribution", {}).keys():
                self._max_core = max(self._max_core, int(core_key))

        core_cols = [f"core_{i}" for i in range(self._max_core + 1)]

        with self._write_lock:
            wb = load_workbook(self.xlsx_path)
            ws = wb["thread_cpu"]
            if not self._header_written:
                ws.append(["timestamp", "tid", "name", "avg_cpu", "min_cpu", "max_cpu",
                            "primary_core"] + core_cols)
                self._header_written = True
            for t in threads:
                core_dist = t.get("core_distribution", {})
                total_samples = sum(core_dist.values()) or 1
                core_pcts = [
                    round(core_dist.get(i, 0) / total_samples * 100, 1)
                    for i in range(self._max_core + 1)
                ]
                ws.append([
                    ts,
                    t["tid"],
                    t["name"],
                    t["cpu_usage"]["avg"],
                    t["cpu_usage"]["min"],
                    t["cpu_usage"]["max"],
                    t.get("primary_core", ""),
                ] + core_pcts)
            wb.save(self.xlsx_path)

    def _run(self):
        while not self._stop_event.is_set():
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            collector = ThreadCPUCollector(pgrep=self._pgrep)
            data = collector.gather(self._sample_interval, include_timeseries=False)
            if "threads" in data:
                try:
                    self._append_batch(ts, data["threads"])
                    with self._lock:
                        self.rows_written += len(data["threads"])
                except Exception as e:
                    print(f"[ThreadMonitor] write error: {e}")


# Global singleton
_thread_monitor = ThreadBackgroundMonitor()


# ============================================================================
# THREAD PLOT HELPERS
# ============================================================================

def _build_affinity_heatmap(xlsx_path: str, label: str,
                             start: Optional[str], end: Optional[str]) -> io.BytesIO:
    """Thread-to-Core Affinity heatmap from Excel"""
    df = pd.read_excel(xlsx_path, sheet_name="thread_cpu", engine="openpyxl")
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    if start:
        df = df[df["timestamp"] >= pd.to_datetime(start)]
    if end:
        df = df[df["timestamp"] <= pd.to_datetime(end)]
    if df.empty:
        raise ValueError("No data in the specified time range")

    core_cols = sorted([c for c in df.columns if re.match(r"^core_\d+$", c)],
                       key=lambda x: int(x.split("_")[1]))

    # Average core % per thread across all samples
    agg = df.groupby("name")[core_cols].mean().round(1)

    # Sort rows by primary_core then name
    primary = df.groupby("name")["primary_core"].agg(lambda x: x.mode()[0])
    agg["_primary"] = primary
    agg = agg.sort_values("_primary").drop(columns=["_primary"])
    agg.columns = [f"CPU{c.split('_')[1]}" for c in agg.columns]

    offline_cpus = []  # could be extended to read from Excel metadata
    fig_h = max(6, len(agg) * 0.45)
    fig, ax = plt.subplots(figsize=(max(14, len(agg.columns) * 0.6), fig_h))
    sns.heatmap(agg, annot=True, fmt=".0f", cmap="YlOrRd",
                cbar_kws={"label": "% of samples"}, ax=ax)

    # Mark offline CPUs
    for cpu_id in offline_cpus:
        col_idx = cpu_id
        if col_idx < len(agg.columns):
            ax.add_patch(plt.Rectangle((col_idx, 0), 1, len(agg),
                                       fill=False, edgecolor="blue",
                                       hatch="///", linewidth=2, zorder=10))

    ax.set_title(f"Thread-to-Core Affinity\n{label}")
    ax.set_xlabel("CPU Core")
    ax.set_ylabel("Thread Name")
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _build_thread_cpu_heatmap(xlsx_path: str, label: str,
                               start: Optional[str], end: Optional[str]) -> io.BytesIO:
    """Thread CPU usage heatmap (min/avg/max) from Excel"""
    df = pd.read_excel(xlsx_path, sheet_name="thread_cpu", engine="openpyxl")
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    if start:
        df = df[df["timestamp"] >= pd.to_datetime(start)]
    if end:
        df = df[df["timestamp"] <= pd.to_datetime(end)]
    if df.empty:
        raise ValueError("No data in the specified time range")

    agg = df.groupby("name").agg(
        min_cpu=("min_cpu", "min"),
        avg_cpu=("avg_cpu", "mean"),
        max_cpu=("max_cpu", "max"),
    ).round(1)
    agg.columns = ["min %", "avg %", "max %"]
    agg = agg.sort_values("avg %", ascending=False)

    fig_h = max(6, len(agg) * 0.4)
    fig, ax = plt.subplots(figsize=(8, fig_h))
    sns.heatmap(agg, annot=True, fmt=".1f", cmap="YlOrRd",
                cbar_kws={"label": "CPU %"}, ax=ax)
    ax.set_title(f"Thread CPU Usage\n{label}")
    ax.set_xlabel("Metric")
    ax.set_ylabel("Thread Name")
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf

# ============================================================================
# BACKGROUND MONITOR — writes to Excel row-by-row
# ============================================================================

class BackgroundMonitor:
    """
    Runs a background thread that samples CPU every second and appends
    one row per sample to an Excel file.

    Excel layout (sheet: "cpu_usage"):
        timestamp        | cpu0 | cpu1 | ... | cpuN
        2026-05-04 10:00:01 | 4.5  | 12.3 | ...
    """

    def __init__(self):
        self._thread: Optional[threading.Thread] = None
        self._stop_event = threading.Event()
        self._lock = threading.Lock()
        self.xlsx_path: Optional[str] = None
        self.running = False
        self.rows_written = 0
        self.started_at: Optional[str] = None
        self._col_order: Optional[List[str]] = None
        self._collector: Optional[CPUCollector] = None
        self._write_lock = threading.Lock()   # separate lock just for file writes

    # ------------------------------------------------------------------ #

    def start(self, xlsx_path: str) -> Dict[str, Any]:
        with self._lock:
            if self.running:
                return {"status": "already_running", "xlsx": self.xlsx_path}

            self.xlsx_path = xlsx_path
            self._stop_event.clear()
            self.rows_written = 0
            self.started_at = datetime.now().isoformat(timespec="seconds")
            self._col_order = None
            self._collector = CPUCollector()

            os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
            self._init_workbook(xlsx_path)

            self._thread = threading.Thread(target=self._run, daemon=True)
            self._thread.start()
            self.running = True

        return {"status": "started", "xlsx": xlsx_path, "started_at": self.started_at}

    def stop(self) -> Dict[str, Any]:
        with self._lock:
            if not self.running:
                return {"status": "not_running"}
            self._stop_event.set()

        if self._thread:
            self._thread.join(timeout=5)

        with self._lock:
            self.running = False

        return {
            "status": "stopped",
            "xlsx": self.xlsx_path,
            "rows_written": self.rows_written,
            "started_at": self.started_at,
            "stopped_at": datetime.now().isoformat(timespec="seconds"),
        }

    def status(self) -> Dict[str, Any]:
        with self._lock:
            return {
                "running": self.running,
                "xlsx": self.xlsx_path,
                "rows_written": self.rows_written,
                "started_at": self.started_at,
            }

    # ------------------------------------------------------------------ #

    def _init_workbook(self, path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "cpu_usage"
        wb.save(path)
        self._header_written = False

    @staticmethod
    def _sort_cpu_keys(keys: List[str]) -> List[str]:
        def _k(k):
            m = re.match(r"^cpu(\d+)$", k)
            return (0, int(m.group(1))) if m else (1, k)
        return sorted(keys, key=_k)

    def _append_row(self, timestamp: str, sample: Dict[str, Any]):
        if self._col_order is None:
            core_keys = [k for k in sample if re.match(r"^cpu\d+$", k)]
            self._col_order = self._sort_cpu_keys(core_keys)

        row = [timestamp] + [
            sample.get(cpu, {}).get("usage", None) for cpu in self._col_order
        ]

        with self._write_lock:
            wb = load_workbook(self.xlsx_path)
            ws = wb["cpu_usage"]
            # Write header exactly once, before first data row
            if not self._header_written:
                ws.append(["timestamp"] + self._col_order)
                self._header_written = True
            ws.append(row)
            wb.save(self.xlsx_path)

    def _run(self):
        # Prime collector — first call builds prev_stats, returns nothing
        self._collector.collect_sample()
        time.sleep(1)

        while not self._stop_event.is_set():
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sample = self._collector.collect_sample()
            if sample:
                try:
                    self._append_row(ts, sample)
                    with self._lock:
                        self.rows_written += 1
                except Exception as e:
                    print(f"[BackgroundMonitor] write error: {e}")
            time.sleep(1)


# Global singleton
_monitor = BackgroundMonitor()


# ============================================================================
# PLOT HELPERS — read from Excel
# ============================================================================

def _sort_cpu_keys(keys: List[str]) -> List[str]:
    def _k(k):
        m = re.match(r"^cpu(\d+)$", k)
        return (0, int(m.group(1))) if m else (1, k)
    return sorted(keys, key=_k)


def _load_excel(xlsx_path: str, start: Optional[str], end: Optional[str]) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name="cpu_usage", engine="openpyxl")
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    if start:
        df = df[df["timestamp"] >= pd.to_datetime(start)]
    if end:
        df = df[df["timestamp"] <= pd.to_datetime(end)]
    if df.empty:
        raise ValueError("No data in the specified time range")
    return df


def _build_heatmap_from_excel(df: pd.DataFrame, label: str) -> io.BytesIO:
    cpu_cols = _sort_cpu_keys([c for c in df.columns if re.match(r"^cpu\d+$", c)])
    stats = pd.DataFrame({
        "min %": df[cpu_cols].min(),
        "avg %": df[cpu_cols].mean().round(2),
        "max %": df[cpu_cols].max(),
    }).loc[cpu_cols]

    fig_h = max(6, len(cpu_cols) * 0.35)
    fig, ax = plt.subplots(figsize=(8, fig_h))
    sns.heatmap(stats, annot=True, fmt=".1f", cmap="YlOrRd",
                cbar_kws={"label": "CPU %"}, ax=ax)
    ax.set_title(f"CPU Usage Heatmap — {label}")
    ax.set_xlabel("Metric")
    ax.set_ylabel("CPU Core")
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _build_timeseries_from_excel(df: pd.DataFrame, label: str) -> io.BytesIO:
    cpu_cols = _sort_cpu_keys([c for c in df.columns if re.match(r"^cpu\d+$", c)])
    t0 = df["timestamp"].iloc[0]
    elapsed = (df["timestamp"] - t0).dt.total_seconds().tolist()

    fig, ax = plt.subplots(figsize=(14, 6))
    cmap = plt.get_cmap("tab20")
    for i, col in enumerate(cpu_cols):
        ax.plot(elapsed, df[col].tolist(), label=col,
                color=cmap(i % 20), linewidth=1)

    ax.set_title(f"CPU Usage Timeseries — {label}")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("CPU Usage (%)")
    ax.set_ylim(0, 100)
    ax.legend(loc="upper right", fontsize=7, ncol=4)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _build_combined_from_excel(df: pd.DataFrame, label: str) -> io.BytesIO:
    cpu_cols = _sort_cpu_keys([c for c in df.columns if re.match(r"^cpu\d+$", c)])
    stats = pd.DataFrame({
        "min %": df[cpu_cols].min(),
        "avg %": df[cpu_cols].mean().round(2),
        "max %": df[cpu_cols].max(),
    }).loc[cpu_cols]

    t0 = df["timestamp"].iloc[0]
    elapsed = (df["timestamp"] - t0).dt.total_seconds().tolist()

    fig_h = max(8, len(cpu_cols) * 0.35)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(22, fig_h))

    sns.heatmap(stats, annot=True, fmt=".1f", cmap="YlOrRd",
                cbar_kws={"label": "CPU %"}, ax=ax1)
    ax1.set_title(f"CPU Usage Heatmap — {label}")
    ax1.set_xlabel("Metric")
    ax1.set_ylabel("CPU Core")

    cmap = plt.get_cmap("tab20")
    for i, col in enumerate(cpu_cols):
        ax2.plot(elapsed, df[col].tolist(), label=col,
                 color=cmap(i % 20), linewidth=1)
    ax2.set_title(f"CPU Usage Timeseries — {label}")
    ax2.set_xlabel("Time (s)")
    ax2.set_ylabel("CPU Usage (%)")
    ax2.set_ylim(0, 100)
    ax2.legend(loc="upper right", fontsize=7, ncol=4)
    ax2.grid(True, alpha=0.3)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


# ============================================================================
# ROUTES
# ============================================================================

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


# --- one-shot monitor (kept for compatibility) ----------------------------

@app.route("/cpu/monitor", methods=["POST"])
def cpu_monitor():
    """Blocking one-shot monitor"""
    data = request.json or {}
    collector = CPUCollector(include_breakdown=data.get("breakdown", False))
    result = collector.gather(
        data.get("duration", 10),
        include_timeseries=data.get("include_timeseries", True),
    )
    return jsonify(result)


# --- background monitor ---------------------------------------------------

@app.route("/cpu/monitor/start", methods=["POST"])
def cpu_monitor_start():
    """
    Start continuous background monitoring, writing every second to Excel.
    If no xlsx path given, auto-generates one with timestamp: /app/cpu_log_YYYYMMDD_HHMMSS.xlsx

    POST body:
    {
        "xlsx": "/app/cpu_log.xlsx"    # optional, auto-named if omitted
    }
    """
    body = request.json or {}
    xlsx_path = body.get("xlsx")
    if not xlsx_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = f"/app/cpu_log_{ts}.xlsx"
    return jsonify(_monitor.start(xlsx_path))


@app.route("/cpu/monitor/stop", methods=["POST"])
def cpu_monitor_stop():
    """Stop background monitoring"""
    return jsonify(_monitor.stop())


@app.route("/cpu/monitor/status", methods=["GET"])
def cpu_monitor_status():
    """Get background monitor status"""
    return jsonify(_monitor.status())


# --- plot from Excel -------------------------------------------------------

@app.route("/cpu/plot", methods=["POST"])
def cpu_plot():
    """
    Generate CPU usage plots from Excel data.
    Saves heatmap and timeseries as separate files with timestamp in filename.

    POST body:
    {
        "xlsx":  "/app/cpu_log_20260504_100000.xlsx",  # required
        "label": "my run",                             # optional title suffix
        "start": "2026-05-04 10:00:00",                # optional time filter
        "end":   "2026-05-04 12:00:00"                 # optional time filter
    }

    Returns: JSON with saved file paths
    {
        "heatmap":    "/app/cpu_plot_heatmap_20260504_102300.png",
        "timeseries": "/app/cpu_plot_timeseries_20260504_102300.png"
    }
    """
    body = request.json or {}
    xlsx_path = body.get("xlsx")
    if not xlsx_path:
        return jsonify({"error": "missing 'xlsx' field"}), 400
    if not os.path.exists(xlsx_path):
        return jsonify({"error": f"file not found: {xlsx_path}"}), 404

    label = body.get("label") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    start = body.get("start")
    end = body.get("end")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    try:
        df = _load_excel(xlsx_path, start, end)
        heatmap_buf    = _build_heatmap_from_excel(df, label)
        timeseries_buf = _build_timeseries_from_excel(df, label)
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        return jsonify({"error": f"plot failed: {e}"}), 500

    saved = {}
    for name, buf in [("heatmap", heatmap_buf), ("timeseries", timeseries_buf)]:
        path = os.path.join("/app", f"cpu_plot_{name}_{ts}.png")
        try:
            with open(path, "wb") as f:
                f.write(buf.getvalue())
            saved[name] = path
        except Exception as e:
            print(f"[cpu_plot] failed to save {name}: {e}")
            saved[name] = f"error: {e}"

    return jsonify(saved)


# --- other endpoints (unchanged) ------------------------------------------

@app.route("/cpu/governor", methods=["GET"])
def cpu_governor():
    return jsonify(CPUGovernorCollector.get_governor())


@app.route("/cpu/idle_states", methods=["GET"])
def cpu_idle_states():
    return jsonify(CPUIdleCollector.get_idle_states())


@app.route("/cpu/context_switches", methods=["POST"])
def context_switches_monitor():
    data = request.json or {}
    collector = ContextSwitchCollector(cpu_filter=data.get("cpu_filter"))
    result = collector.gather(data.get("duration", 10))
    return jsonify(result)


@app.route("/irq/affinity", methods=["GET"])
def irq_affinity():
    return jsonify(IRQAffinityCollector.get_affinity(request.args.get("pattern", "ens")))


# ============================================================================
# THREAD ROUTES
# ============================================================================

@app.route("/thread/monitor/start", methods=["POST"])
def thread_monitor_start():
    """
    Start continuous background thread monitoring, writing to Excel.

    POST body:
    {
        "xlsx":            "/app/thread_log_20260504.xlsx",  # optional, auto-named if omitted
        "pgrep":           "nr-softmodem",                   # process to monitor (default: nr-softmodem)
        "sample_interval": 10                                # seconds per sample (default: 10)
    }
    """
    body = request.json or {}
    xlsx_path = body.get("xlsx")
    if not xlsx_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = f"/app/thread_log_{ts}.xlsx"
    pgrep = body.get("pgrep", "nr-softmodem")
    sample_interval = body.get("sample_interval", 10)
    return jsonify(_thread_monitor.start(xlsx_path, pgrep=pgrep, sample_interval=sample_interval))


@app.route("/thread/monitor/stop", methods=["POST"])
def thread_monitor_stop():
    """Stop background thread monitoring"""
    return jsonify(_thread_monitor.stop())


@app.route("/thread/monitor/status", methods=["GET"])
def thread_monitor_status():
    """Get background thread monitor status"""
    return jsonify(_thread_monitor.status())


@app.route("/thread/plot", methods=["POST"])
def thread_plot():
    """
    Generate thread plots from Excel data.
    Saves both affinity heatmap and CPU usage heatmap as separate files.

    POST body:
    {
        "xlsx":  "/app/thread_log_20260504_100000.xlsx",  # required
        "label": "my run",                                # optional title suffix
        "start": "2026-05-04 10:00:00",                   # optional time filter
        "end":   "2026-05-04 12:00:00"                    # optional time filter
    }

    Returns JSON:
    {
        "affinity":  "/app/thread_plot_affinity_20260504_102300.png",
        "cpu_usage": "/app/thread_plot_cpu_usage_20260504_102300.png"
    }
    """
    body = request.json or {}
    xlsx_path = body.get("xlsx")
    if not xlsx_path:
        return jsonify({"error": "missing 'xlsx' field"}), 400
    if not os.path.exists(xlsx_path):
        return jsonify({"error": f"file not found: {xlsx_path}"}), 404

    label = body.get("label") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    start = body.get("start")
    end = body.get("end")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    saved = {}
    plots = {
        "affinity":  _build_affinity_heatmap,
        "cpu_usage": _build_thread_cpu_heatmap,
    }
    for name, fn in plots.items():
        try:
            buf = fn(xlsx_path, label, start, end)
            path = os.path.join("/app", f"thread_plot_{name}_{ts}.png")
            with open(path, "wb") as f:
                f.write(buf.getvalue())
            saved[name] = path
        except ValueError as e:
            return jsonify({"error": str(e)}), 422
        except Exception as e:
            saved[name] = f"error: {e}"

    return jsonify(saved)


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    # debug=False for long-running stability
    app.run(host="0.0.0.0", port=PORT, debug=False, use_reloader=False)
